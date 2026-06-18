import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
import subprocess
import stat
import sys
import queue
import threading
import datetime
import re
import customtkinter as ctk
from tkinter import filedialog, messagebox
import webbrowser
import io

log_queue = queue.Queue()
progress_queue = queue.Queue()

PROJECT_URL = "https://github.com/caifugao110/heyanlin/tree/master/table-comparison"

DEFAULT_APPEARANCE_MODE = "light"
DEFAULT_COLOR_THEME = "blue"

ctk.set_appearance_mode(DEFAULT_APPEARANCE_MODE)
ctk.set_default_color_theme(DEFAULT_COLOR_THEME)


def bundled_path(name: str) -> str:
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, name)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), name)


def load_project_metadata() -> dict[str, str]:
    pyproject_path = bundled_path("pyproject.toml")
    metadata = {"version": "unknown", "author": "unknown", "homepage": ""}
    if not os.path.exists(pyproject_path):
        return metadata

    try:
        with open(pyproject_path, "r", encoding="utf-8") as f:
            text = f.read()
    except Exception:
        return metadata

    try:
        import tomllib
        project = tomllib.loads(text).get("project", {})
        urls = project.get("urls", {})
        authors = project.get("authors", [])
        metadata["version"] = project.get("version", metadata["version"])
        if authors:
            metadata["author"] = authors[0].get("name", metadata["author"])
        metadata["homepage"] = urls.get("Homepage", metadata["homepage"])
        return metadata
    except Exception:
        pass

    patterns = {
        "version": r'(?m)^version\s*=\s*"([^"]+)"',
        "author": r'authors\s*=\s*\[\{\s*name\s*=\s*"([^"]+)"',
        "homepage": r'(?m)^Homepage\s*=\s*"([^"]+)"',
    }
    for key, pattern in patterns.items():
        match = re.search(pattern, text)
        if match:
            metadata[key] = match.group(1)
    return metadata


PROJECT_METADATA = load_project_metadata()
__version__ = PROJECT_METADATA["version"]
__author__ = PROJECT_METADATA["author"]
__homepage__ = PROJECT_METADATA["homepage"]


def cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def compare_excel_files(baseline_path, compare_path, output_baseline_path, output_compare_path, results_folder, original_filename, timestamp, header_row=3, key_fields=None, stop_event=None):
    def check_stop():
        if stop_event and stop_event.is_set():
            log_queue.put("操作已取消")
            return True
        return False

    fill_changed = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_added = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    fill_deleted = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    try:
        if check_stop():
            return False

        wb_baseline = openpyxl.load_workbook(baseline_path, data_only=True)
        wb_compare = openpyxl.load_workbook(compare_path, data_only=True)
    except FileNotFoundError as e:
        log_queue.put(f"错误：找不到文件 - {e}")
        return False
    except Exception as e:
        log_queue.put(f"加载文件时出错: {e}")
        return False

    ws_baseline = wb_baseline.active
    ws_compare = wb_compare.active

    baseline_max_row = ws_baseline.max_row
    baseline_max_col = ws_baseline.max_column
    compare_max_row = ws_compare.max_row
    compare_max_col = ws_compare.max_column

    if baseline_max_col != compare_max_col:
        log_queue.put(f"警告：两个文件的列数不一致！基准文件：{baseline_max_col}列，比较文件：{compare_max_col}列")

    cells_baseline = {}
    cells_compare = {}

    for r in range(1, baseline_max_row + 1):
        if check_stop():
            return False
        for c in range(1, baseline_max_col + 1):
            cells_baseline[(r, c)] = ws_baseline.cell(row=r, column=c).value

    for r in range(1, compare_max_row + 1):
        if check_stop():
            return False
        for c in range(1, compare_max_col + 1):
            cells_compare[(r, c)] = ws_compare.cell(row=r, column=c).value

    if not key_fields:
        header_values = [cell_text(cells_baseline.get((header_row, c))) for c in range(1, min(baseline_max_col + 1, 4))]
        key_fields = [v for v in header_values if v]
        if len(key_fields) < 3:
            key_fields = [f"列{c}" for c in range(1, min(baseline_max_col + 1, 4))]

    def find_key_columns(cells, max_col, header_row_num, key_field_names):
        key_cols = {}
        header_values = {}
        for col in range(1, max_col + 1):
            cell_value = cell_text(cells.get((header_row_num, col)))
            header_values[cell_value] = col

        for field in key_field_names:
            if field in header_values:
                key_cols[field] = header_values[field]
            else:
                try:
                    col_idx = int(field.replace("列", ""))
                    if 1 <= col_idx <= max_col:
                        key_cols[field] = col_idx
                except ValueError:
                    pass
        return key_cols

    key_cols_baseline = find_key_columns(cells_baseline, baseline_max_col, header_row, key_fields)
    key_cols_compare = find_key_columns(cells_compare, compare_max_col, header_row, key_fields)

    has_all_keys_baseline = all(field in key_cols_baseline for field in key_fields)
    has_all_keys_compare = all(field in key_cols_compare for field in key_fields)

    row_mapping = {}

    if has_all_keys_baseline and has_all_keys_compare:
        def build_row_key_map(cells, max_row, key_cols, data_start_row):
            row_key_map = {}
            for row in range(data_start_row, max_row + 1):
                key_values = tuple(cells.get((row, key_cols[field]), None) for field in key_fields)
                if all(v is not None for v in key_values):
                    row_key_map[key_values] = row
            return row_key_map

        data_start_row = header_row + 1
        row_key_map_baseline = build_row_key_map(cells_baseline, baseline_max_row, key_cols_baseline, data_start_row)
        row_key_map_compare = build_row_key_map(cells_compare, compare_max_row, key_cols_compare, data_start_row)

        for key in row_key_map_baseline:
            if key in row_key_map_compare:
                row_baseline = row_key_map_baseline[key]
                row_compare = row_key_map_compare[key]
                row_mapping[row_baseline] = row_compare
    else:
        log_queue.put("\n无法找到所有关键字段，使用默认行匹配...")

        def get_row_content(row_num, cells, max_col):
            return tuple(cells.get((row_num, c), None) for c in range(1, max_col + 1))

        row_contents_baseline = {r: get_row_content(r, cells_baseline, baseline_max_col) for r in range(1, baseline_max_row + 1)}
        row_contents_compare = {r: get_row_content(r, cells_compare, compare_max_col) for r in range(1, compare_max_row + 1)}

        for row_baseline, content_baseline in row_contents_baseline.items():
            if check_stop():
                return False

            for row_compare, content_compare in row_contents_compare.items():
                if row_compare not in row_mapping.values() and content_baseline == content_compare:
                    row_mapping[row_baseline] = row_compare
                    break

        if len(row_mapping) < min(baseline_max_row, compare_max_row) // 2:
            min_rows = min(baseline_max_row, compare_max_row)
            row_mapping = {r: r for r in range(1, min_rows + 1)}

    changes_count = 0
    added_rows_count = 0
    deleted_rows_count = 0

    key_col_set_baseline = set(key_cols_baseline.values()) if has_all_keys_baseline else set()
    key_col_set_compare = set(key_cols_compare.values()) if has_all_keys_compare else set()

    log_queue.put("\n开始比较匹配行的单元格差异...")

    def create_col_name_map():
        col_name_map = {}
        baseline_col_names = {}
        for col_b in range(1, baseline_max_col + 1):
            col_name_b = cell_text(cells_baseline.get((header_row, col_b)))
            if col_name_b:
                baseline_col_names[col_name_b] = col_b

        for col_c in range(1, compare_max_col + 1):
            col_name_c = cell_text(cells_compare.get((header_row, col_c)))
            if col_name_c in baseline_col_names:
                col_name_map[baseline_col_names[col_name_c]] = col_c

        if len(col_name_map) < min(baseline_max_col, compare_max_col) // 2:
            min_cols = min(baseline_max_col, compare_max_col)
            col_name_map = {c: c for c in range(1, min_cols + 1)}

        return col_name_map

    col_name_map = create_col_name_map()

    for row_baseline, row_compare in row_mapping.items():
        if check_stop():
            return False

        for col_baseline, col_compare in col_name_map.items():
            if col_baseline in key_col_set_baseline or col_compare in key_col_set_compare:
                continue

            val_baseline = cells_baseline.get((row_baseline, col_baseline), None)
            val_compare = cells_compare.get((row_compare, col_compare), None)

            if val_baseline != val_compare:
                ws_baseline.cell(row=row_baseline, column=col_baseline).fill = fill_changed
                ws_compare.cell(row=row_compare, column=col_compare).fill = fill_changed
                changes_count += 1

    log_queue.put("\n开始标记新增行、删除行和数值变化行...")

    def get_all_row_keys(cells, max_row, key_cols, data_start_row):
        all_row_keys = {}
        for row in range(data_start_row, max_row + 1):
            key_values = tuple(cells.get((row, key_cols[field]), None) for field in key_fields)
            if all(v is not None for v in key_values):
                all_row_keys[key_values] = row
        return all_row_keys

    if has_all_keys_baseline and has_all_keys_compare:
        data_start_row = header_row + 1
        all_baseline_keys = get_all_row_keys(cells_baseline, baseline_max_row, key_cols_baseline, data_start_row)
        all_compare_keys = get_all_row_keys(cells_compare, compare_max_row, key_cols_compare, data_start_row)

        for key, row_baseline in all_baseline_keys.items():
            if check_stop():
                return False

            if key not in all_compare_keys:
                for col in range(1, baseline_max_col + 1):
                    ws_baseline.cell(row=row_baseline, column=col).fill = fill_added
                deleted_rows_count += 1
        log_queue.put(f"\n已标记 {deleted_rows_count} 行删除（绿色）")

        for key, row_compare in all_compare_keys.items():
            if check_stop():
                return False

            if key not in all_baseline_keys:
                for col in range(1, compare_max_col + 1):
                    ws_compare.cell(row=row_compare, column=col).fill = fill_deleted
                added_rows_count += 1
        log_queue.put(f"\n已标记 {added_rows_count} 行新增（红色）")
    else:
        log_queue.put("\n使用简单匹配标记新增和删除行...")

        for row_baseline in range(1, baseline_max_row + 1):
            if check_stop():
                return False

            if row_baseline not in row_mapping:
                for col in range(1, baseline_max_col + 1):
                    ws_baseline.cell(row=row_baseline, column=col).fill = fill_added
                deleted_rows_count += 1
        log_queue.put(f"\n已标记 {deleted_rows_count} 行删除（绿色）")

        mapped_compare_rows = set(row_mapping.values())
        for row_compare in range(1, compare_max_row + 1):
            if check_stop():
                return False

            if row_compare not in mapped_compare_rows:
                for col in range(1, compare_max_col + 1):
                    ws_compare.cell(row=row_compare, column=col).fill = fill_deleted
                added_rows_count += 1
        log_queue.put(f"\n已标记 {added_rows_count} 行新增（红色）")

    if changes_count > 0:
        log_queue.put(f"\n已标记 {changes_count} 处数值变化（黄色）")

    total_changes = changes_count + added_rows_count + deleted_rows_count
    log_queue.put(f"\n比较完成！共发现 {total_changes} 处差异。")

    try:
        wb_baseline.save(output_baseline_path)
        wb_compare.save(output_compare_path)
    except Exception as e:
        log_queue.put(f"保存结果文件时出错: {e}")
        return False

    log_queue.put("\n正在生成差异结果文件...")

    try:
        wb_diff = openpyxl.load_workbook(output_baseline_path)
        ws_diff = wb_diff.active
        ws_diff.title = "差异比较结果"

        wb_baseline_saved = openpyxl.load_workbook(output_baseline_path)
        ws_baseline_saved = wb_baseline_saved.active

        wb_compare_saved = openpyxl.load_workbook(output_compare_path)
        ws_compare_saved = wb_compare_saved.active
    except Exception as e:
        log_queue.put(f"加载保存后的文件时出错: {e}")
        return False

    if has_all_keys_baseline and has_all_keys_compare:
        key_to_row = {}

        data_start_row_diff = header_row + 1
        for row_baseline in range(data_start_row_diff, ws_baseline_saved.max_row + 1):
            if check_stop():
                return False

            key_values = tuple(ws_baseline_saved.cell(row=row_baseline, column=key_cols_baseline[field]).value for field in key_fields)
            if all(v is not None for v in key_values):
                key_to_row[key_values] = row_baseline

        added_rows = []
        for row_compare in range(data_start_row_diff, ws_compare_saved.max_row + 1):
            if check_stop():
                return False

            key_values = tuple(ws_compare_saved.cell(row=row_compare, column=key_cols_compare[field]).value for field in key_fields)
            if not all(v is not None for v in key_values):
                continue

            first_cell = ws_compare_saved.cell(row=row_compare, column=1)
            if first_cell.fill.start_color.rgb == fill_deleted.start_color.rgb:
                prev_key_values = None
                if row_compare > data_start_row_diff:
                    prev_key_values = tuple(ws_compare_saved.cell(row=row_compare - 1, column=key_cols_compare[field]).value for field in key_fields)
                added_rows.append((key_values, row_compare, prev_key_values))

        added_rows_seen = set()
        unique_added_rows = []
        for key_values, row_compare, prev_key_values in added_rows:
            if key_values not in added_rows_seen:
                added_rows_seen.add(key_values)
                unique_added_rows.append((key_values, row_compare, prev_key_values))

        unique_added_rows.sort(key=lambda x: x[1])

        for key_values, row_compare, prev_key_values in unique_added_rows:
            if check_stop():
                return False

            insert_row = ws_diff.max_row
            if prev_key_values and prev_key_values in key_to_row:
                insert_row = key_to_row[prev_key_values] + 1
            elif key_values in key_to_row:
                insert_row = key_to_row[key_values] + 1

            ws_diff.insert_rows(insert_row)

            for k, v in list(key_to_row.items()):
                if v >= insert_row:
                    key_to_row[k] = v + 1

            template_row = data_start_row_diff

            for col in range(1, baseline_max_col + 1):
                template_cell = ws_baseline_saved.cell(row=template_row, column=col)
                new_cell = ws_diff.cell(row=insert_row, column=col)

                new_cell.number_format = template_cell.number_format
                new_cell.font = Font(**template_cell.font.__dict__)
                new_cell.border = Border(**template_cell.border.__dict__)
                new_cell.alignment = Alignment(**template_cell.alignment.__dict__)

            for col in range(1, baseline_max_col + 1):
                col_name_b = ws_baseline_saved.cell(row=header_row, column=col).value
                col_name_b = cell_text(col_name_b)
                if not col_name_b:
                    continue

                for c in range(1, ws_compare_saved.max_column + 1):
                    col_name_c = ws_compare_saved.cell(row=header_row, column=c).value
                    col_name_c = cell_text(col_name_c)
                    if col_name_c == col_name_b:
                        value = ws_compare_saved.cell(row=row_compare, column=c).value
                        ws_diff.cell(row=insert_row, column=col, value=value)
                        break

            for col in range(1, baseline_max_col + 1):
                ws_diff.cell(row=insert_row, column=col).fill = fill_deleted
    else:
        data_start_row_diff = header_row + 1
        added_rows = []
        mapped_compare_rows = set(row_mapping.values())

        for row_compare in range(data_start_row_diff, ws_compare_saved.max_row + 1):
            if check_stop():
                return False

            if row_compare not in mapped_compare_rows:
                first_cell = ws_compare_saved.cell(row=row_compare, column=1)
                if first_cell.fill.start_color.rgb == fill_deleted.start_color.rgb:
                    added_rows.append(row_compare)

        added_rows.sort()

        for row_compare in added_rows:
            if check_stop():
                return False

            insert_row = ws_diff.max_row

            ws_diff.insert_rows(insert_row)

            template_row = data_start_row_diff

            for col in range(1, baseline_max_col + 1):
                template_cell = ws_baseline_saved.cell(row=template_row, column=col)
                new_cell = ws_diff.cell(row=insert_row, column=col)

                new_cell.number_format = template_cell.number_format
                new_cell.font = Font(**template_cell.font.__dict__)
                new_cell.border = Border(**template_cell.border.__dict__)
                new_cell.alignment = Alignment(**template_cell.alignment.__dict__)

            min_cols = min(baseline_max_col, ws_compare_saved.max_column)
            for col in range(1, min_cols + 1):
                value = ws_compare_saved.cell(row=row_compare, column=col).value
                ws_diff.cell(row=insert_row, column=col, value=value)

            for col in range(1, baseline_max_col + 1):
                ws_diff.cell(row=insert_row, column=col).fill = fill_deleted

    for col in range(1, ws_baseline_saved.max_column + 1):
        if check_stop():
            return False

        col_letter = get_column_letter(col)
        if col_letter in ws_baseline_saved.column_dimensions:
            ws_diff.column_dimensions[col_letter].width = ws_baseline_saved.column_dimensions[col_letter].width

    for row in range(1, ws_baseline_saved.max_row + 1):
        if check_stop():
            return False

        if row in ws_baseline_saved.row_dimensions:
            ws_diff.row_dimensions[row].height = ws_baseline_saved.row_dimensions[row].height

    diff_output_path = os.path.join(results_folder, f"差异结果_{timestamp}.xlsx")
    try:
        wb_diff.save(diff_output_path)
    except Exception as e:
        log_queue.put(f"保存差异结果文件时出错: {e}")
        return False

    try:
        baseline_stat = os.stat(output_baseline_path)
        compare_stat = os.stat(output_compare_path)
        diff_stat = os.stat(diff_output_path)

        if os.name == 'nt':
            subprocess.run(['attrib', '+r', output_baseline_path], check=True)
            subprocess.run(['attrib', '+r', output_compare_path], check=True)
            subprocess.run(['attrib', '+r', diff_output_path], check=True)
        else:
            os.chmod(output_baseline_path, baseline_stat.st_mode & ~stat.S_IWUSR & ~stat.S_IWGRP & ~stat.S_IWOTH)
            os.chmod(output_compare_path, compare_stat.st_mode & ~stat.S_IWUSR & ~stat.S_IWGRP & ~stat.S_IWOTH)
            os.chmod(diff_output_path, diff_stat.st_mode & ~stat.S_IWUSR & ~stat.S_IWGRP & ~stat.S_IWOTH)
    except Exception as e:
        log_queue.put(f"设置只读属性时出错: {e}")

    log_queue.put(f"\n已生成差异结果文件至: \n{diff_output_path}")

    try:
        subprocess.Popen(['start', '', output_baseline_path], shell=True)
        subprocess.Popen(['start', '', output_compare_path], shell=True)
        subprocess.Popen(['start', '', diff_output_path], shell=True)
    except Exception as e:
        log_queue.put(f"打开文件时出错: {e}")

    return True


class StdoutRedirector:
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, message):
        log_queue.put(message)

    def flush(self):
        pass


class ExcelCompareGUI(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title(f"Excel文件比较工具 V{__version__}")
        self.geometry("1200x800")
        self.minsize(1000, 700)

        if getattr(sys, "frozen", False):
            self.current_dir = os.path.dirname(sys.executable)
        else:
            self.current_dir = os.path.dirname(os.path.abspath(__file__))
        self.results_folder = os.path.join(self.current_dir, "results")
        os.makedirs(self.results_folder, exist_ok=True)

        self.baseline_file = ""
        self.compare_file = ""
        self.running = False
        self.stop_event = threading.Event()
        self.worker_thread = None

        self._set_icon()

        self._init_widgets()

        self._redirect_stdout()

        self._listen_queues()

    def _set_icon(self):
        try:
            icon_path = bundled_path("assets/app.ico")
            if os.path.exists(icon_path):
                self.wm_iconbitmap(icon_path)
            else:
                raise FileNotFoundError(f"图标文件不存在: {icon_path}")
        except Exception as e:
            print(f"设置主窗口图标失败: {e}")
            try:
                from PIL import Image, ImageTk
                icon_path = bundled_path("assets/app.ico")
                if os.path.exists(icon_path):
                    icon = Image.open(icon_path)
                    icon = icon.resize((32, 32), Image.LANCZOS)
                    self.app_icon = ImageTk.PhotoImage(icon)
                    self.iconphoto(False, self.app_icon)
            except Exception as e2:
                print(f"使用PIL设置图标也失败: {e2}")

    def _init_widgets(self):
        main_container = ctk.CTkFrame(self)
        main_container.pack(fill="both", expand=True, padx=0, pady=0)

        header_frame = ctk.CTkFrame(main_container, fg_color=("gray90", "gray20"), height=100)
        header_frame.pack(fill="x", padx=0, pady=0)
        header_frame.pack_propagate(False)

        title_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        title_frame.pack(fill="x", padx=20, pady=10)

        title_label = ctk.CTkLabel(
            title_frame,
            text="Excel文件比较工具",
            font=("微软雅黑", 26, "bold"),
            text_color=("#1f77b4", "#64b5f6")
        )
        title_label.pack(anchor="w", side="left")

        theme_frame = ctk.CTkFrame(title_frame, fg_color="transparent")
        theme_frame.pack(anchor="e", side="right")

        ctk.CTkLabel(
            theme_frame,
            text="主题:",
            font=("微软雅黑", 12),
            text_color=("gray50", "gray70")
        ).pack(side="left", padx=(0, 10))

        self.appearance_mode_optionemenu = ctk.CTkOptionMenu(
            theme_frame,
            values=["light", "dark", "system"],
            command=self._change_appearance_mode_event,
            font=("微软雅黑", 12),
            width=120
        )
        self.appearance_mode_optionemenu.set(DEFAULT_APPEARANCE_MODE)
        self.appearance_mode_optionemenu.pack(side="left", padx=(0, 10))

        self.color_theme_optionemenu = ctk.CTkOptionMenu(
            theme_frame,
            values=["blue", "green", "dark-blue"],
            command=self._change_color_theme_event,
            font=("微软雅黑", 12),
            width=120
        )
        self.color_theme_optionemenu.set(DEFAULT_COLOR_THEME)
        self.color_theme_optionemenu.pack(side="left")

        info_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        info_frame.pack(anchor="w", padx=20, pady=(0, 10))

        self.version_label = ctk.CTkLabel(
            info_frame,
            text=f"{__author__} © 2026 | V{__version__}",
            font=("微软雅黑", 12),
            text_color=("gray50", "gray70")
        )
        self.version_label.pack(side="left", padx=(0, 20))

        github_btn = ctk.CTkButton(
            info_frame,
            text="📌 GitHub地址",
            width=120,
            height=30,
            font=("微软雅黑", 12),
            command=lambda: webbrowser.open(__homepage__)
        )
        github_btn.pack(side="left", padx=5)

        help_btn = ctk.CTkButton(
            info_frame,
            text="❓ 使用说明",
            width=120,
            height=30,
            font=("微软雅黑", 12),
            command=lambda: webbrowser.open("https://github.com/caifugao110/heyanlin/blob/master/table-comparison/README.md")
        )
        help_btn.pack(side="left", padx=5)

        content_frame = ctk.CTkFrame(main_container, fg_color="transparent")
        content_frame.pack(fill="both", expand=True, padx=15, pady=15)

        left_panel = ctk.CTkFrame(content_frame, fg_color=("gray86", "gray17"))
        left_panel.pack(side="left", fill="y", expand=False, padx=(0, 10))
        left_panel.configure(width=300)

        file_section = ctk.CTkFrame(left_panel, fg_color="transparent")
        file_section.pack(fill="x", padx=15, pady=15)

        ctk.CTkLabel(
            file_section,
            text="文件选择",
            font=("微软雅黑", 16, "bold")
        ).pack(anchor="w", pady=(0, 10))

        baseline_frame = ctk.CTkFrame(file_section, fg_color="transparent")
        baseline_frame.pack(fill="x", pady=5)

        ctk.CTkLabel(
            baseline_frame,
            text="基准文件:",
            width=100,
            font=("微软雅黑", 12)
        ).pack(side="left", anchor="center")

        self.baseline_entry = ctk.CTkEntry(baseline_frame, font=("微软雅黑", 12))
        self.baseline_entry.pack(side="left", fill="x", expand=True, padx=5)

        ctk.CTkButton(
            baseline_frame,
            text="浏览",
            width=60,
            font=("微软雅黑", 12),
            command=self._browse_baseline_file
        ).pack(side="left", padx=5)

        ctk.CTkLabel(
            file_section,
            text="通常选择改动前的文件或者原始文件，该文件内容的新增标记为绿色",
            font=("微软雅黑", 10),
            text_color="gray50"
        ).pack(anchor="w", pady=(0, 5))

        compare_frame = ctk.CTkFrame(file_section, fg_color="transparent")
        compare_frame.pack(fill="x", pady=5)

        ctk.CTkLabel(
            compare_frame,
            text="比较文件:",
            width=100,
            font=("微软雅黑", 12)
        ).pack(side="left", anchor="center")

        self.compare_entry = ctk.CTkEntry(compare_frame, font=("微软雅黑", 12))
        self.compare_entry.pack(side="left", fill="x", expand=True, padx=5)

        ctk.CTkButton(
            compare_frame,
            text="浏览",
            width=60,
            font=("微软雅黑", 12),
            command=self._browse_compare_file
        ).pack(side="left", padx=5)

        ctk.CTkLabel(
            file_section,
            text="通常选择改动后的文件或者新的文件，该文件内容的新增标记为红色",
            font=("微软雅黑", 10),
            text_color="gray50"
        ).pack(anchor="w", pady=(0, 5))

        config_section = ctk.CTkFrame(left_panel, fg_color="transparent")
        config_section.pack(fill="x", padx=15, pady=15)

        ctk.CTkLabel(
            config_section,
            text="比较配置",
            font=("微软雅黑", 16, "bold")
        ).pack(anchor="w", pady=(0, 10))

        header_row_frame = ctk.CTkFrame(config_section, fg_color="transparent")
        header_row_frame.pack(fill="x", pady=5)

        ctk.CTkLabel(
            header_row_frame,
            text="表头行号:",
            width=100,
            font=("微软雅黑", 12)
        ).pack(side="left", anchor="center")

        self.header_row_var = ctk.StringVar(value="")
        self.header_row_entry = ctk.CTkEntry(header_row_frame, textvariable=self.header_row_var, font=("微软雅黑", 12), width=150, state="readonly")
        self.header_row_entry.pack(side="left", padx=5)

        ctk.CTkButton(
            header_row_frame,
            text="选择",
            width=60,
            font=("微软雅黑", 12),
            command=self._select_header_row
        ).pack(side="left", padx=5)

        self.header_preview_frame = ctk.CTkFrame(config_section, fg_color="transparent")
        self.header_preview_frame.pack(fill="x", pady=5)

        self.header_preview_label = ctk.CTkLabel(
            self.header_preview_frame,
            text="请点击'选择'按钮查看并选择表头行号",
            font=("微软雅黑", 10),
            text_color="gray50"
        )
        self.header_preview_label.pack(anchor="w")

        feature_cols_frame = ctk.CTkFrame(config_section, fg_color="transparent")
        feature_cols_frame.pack(fill="x", pady=5)

        ctk.CTkLabel(
            feature_cols_frame,
            text="特征列:",
            width=100,
            font=("微软雅黑", 12)
        ).pack(side="left", anchor="center")

        self.feature_cols_var = ctk.StringVar(value="1,2,3")
        self.feature_cols_entry = ctk.CTkEntry(feature_cols_frame, textvariable=self.feature_cols_var, font=("微软雅黑", 12), width=150, state="readonly")
        self.feature_cols_entry.pack(side="left", padx=5)

        ctk.CTkButton(
            feature_cols_frame,
            text="选择",
            width=60,
            font=("微软雅黑", 12),
            command=self._select_feature_columns
        ).pack(side="left", padx=5)

        self.feature_cols_preview_frame = ctk.CTkFrame(config_section, fg_color="transparent")
        self.feature_cols_preview_frame.pack(fill="x", pady=5)

        self.feature_cols_preview_label = ctk.CTkLabel(
            self.feature_cols_preview_frame,
            text="请点击'选择'按钮查看并选择特征列，最多支持6列，默认使用列: 1,2,3",
            font=("微软雅黑", 10),
            text_color="gray50"
        )
        self.feature_cols_preview_label.pack(anchor="w")

        ctk.CTkLabel(
            config_section,
            text="提示: 特征列用于判断行的增删变化，特征列内容的变化不视为数值变化",
            font=("微软雅黑", 12, "bold"),
            text_color="#FF6B35"
        ).pack(anchor="w", pady=(5, 0))

        button_section = ctk.CTkFrame(left_panel, fg_color="transparent")
        button_section.pack(fill="x", padx=15, pady=15)

        self.start_button = ctk.CTkButton(
            button_section,
            text="开始比较",
            font=("微软雅黑", 16, "bold"),
            height=50,
            fg_color="#4CAF50",
            hover_color="#45a049",
            command=self._start_compare
        )
        self.start_button.pack(fill="x", pady=5)

        self.stop_button = ctk.CTkButton(
            button_section,
            text="停止",
            font=("微软雅黑", 16, "bold"),
            height=50,
            fg_color="#f44336",
            hover_color="#da190b",
            command=self._stop_compare,
            state="disabled"
        )
        self.stop_button.pack(fill="x", pady=5)

        right_panel = ctk.CTkFrame(content_frame, fg_color=("gray86", "gray17"))
        right_panel.pack(side="right", fill="both", expand=True)

        log_title_frame = ctk.CTkFrame(right_panel, fg_color="transparent")
        log_title_frame.pack(fill="x", padx=15, pady=10)

        ctk.CTkLabel(
            log_title_frame,
            text="任务日志",
            font=("微软雅黑", 16, "bold")
        ).pack(anchor="w")

        log_frame = ctk.CTkFrame(right_panel, fg_color="transparent")
        log_frame.pack(fill="both", expand=True, padx=15, pady=5)

        self.log_text = ctk.CTkTextbox(
            log_frame,
            font=("微软雅黑", 12),
            wrap="word",
            corner_radius=8,
            border_width=2,
            border_color=("#D1D1D6", "#4A4A4A"),
            fg_color=("#F8F8F8", "#1A1A1A"),
            text_color=("#424242", "#B0BEC5"),
            padx=10,
            pady=10,
            height=80
        )
        log_frame.grid_rowconfigure(0, weight=1)
        log_frame.grid_columnconfigure(0, weight=1)
        self.log_text.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)

        scrollbar = ctk.CTkScrollbar(
            log_frame,
            command=self.log_text.yview,
            corner_radius=8
        )
        scrollbar.grid(row=0, column=1, sticky="ns", padx=(0, 5), pady=5)
        self.log_text.configure(yscrollcommand=scrollbar.set)

    def _change_appearance_mode_event(self, new_appearance_mode: str):
        ctk.set_appearance_mode(new_appearance_mode)

    def _change_color_theme_event(self, new_color_theme: str):
        ctk.set_default_color_theme(new_color_theme)

    def _browse_baseline_file(self):
        file_path = filedialog.askopenfilename(
            title="选择基准Excel文件",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*")]
        )
        if file_path:
            self.baseline_entry.delete(0, ctk.END)
            self.baseline_entry.insert(0, file_path)
            self.baseline_file = file_path

    def _browse_compare_file(self):
        file_path = filedialog.askopenfilename(
            title="选择比较Excel文件",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*")]
        )
        if file_path:
            self.compare_entry.delete(0, ctk.END)
            self.compare_entry.insert(0, file_path)
            self.compare_file = file_path

    def _start_compare(self):
        self.baseline_file = self.baseline_entry.get().strip()
        self.compare_file = self.compare_entry.get().strip()

        if not self.baseline_file or not self.compare_file:
            messagebox.showerror("错误", "请选择基准文件和比较文件")
            return

        if not os.path.exists(self.baseline_file):
            messagebox.showerror("错误", f"基准文件不存在: {self.baseline_file}")
            return

        if not os.path.exists(self.compare_file):
            messagebox.showerror("错误", f"比较文件不存在: {self.compare_file}")
            return

        header_row_str = self.header_row_var.get().strip()
        if not header_row_str:
            messagebox.showerror("错误", "请选择表头行号")
            return

        self.running = True
        self.stop_event.clear()
        self.start_button.configure(state="disabled")
        self.stop_button.configure(state="normal")

        self.log_text.delete("1.0", ctk.END)

        self.worker_thread = threading.Thread(
            target=self._compare_worker,
            daemon=True
        )
        self.worker_thread.start()

    def _stop_compare(self):
        self.stop_event.set()
        self.stop_button.configure(state="disabled")

    def _select_header_row(self):
        if not self.baseline_file:
            messagebox.showerror("错误", "请先选择基准文件")
            return

        try:
            wb = openpyxl.load_workbook(self.baseline_file, data_only=True)
            ws = wb.active

            max_row = min(10, ws.max_row)
            max_col = min(6, ws.max_column)

            select_window = ctk.CTkToplevel(self)
            select_window.title("选择表头行号")
            select_window.geometry("900x400")

            try:
                icon_path = bundled_path("assets/app.ico")
                if os.path.exists(icon_path):
                    select_window.after(300, lambda: select_window.wm_iconbitmap(icon_path))
            except Exception as e:
                print(f"设置表头行选择窗口图标失败: {e}")

            select_window.transient(self)
            select_window.grab_set()

            preview_frame = ctk.CTkFrame(select_window)
            preview_frame.pack(fill="both", expand=True, padx=10, pady=10)

            for row in range(1, max_row + 1):
                row_btn = ctk.CTkButton(
                    preview_frame,
                    text=f"行 {row}",
                    width=60,
                    height=30,
                    font=("微软雅黑", 10),
                    command=lambda r=row: self._set_header_row(r, select_window)
                )
                row_btn.grid(row=row, column=0, padx=5, pady=2, sticky="w")

                for col in range(1, max_col + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    cell_text = str(cell_value) if cell_value else "空"

                    cell_label = ctk.CTkLabel(
                        preview_frame,
                        text=cell_text,
                        width=140,
                        height=30,
                        font=("微软雅黑", 10),
                        anchor="w"
                    )
                    cell_label.grid(row=row, column=col, padx=5, pady=2, sticky="w")

            info_label = ctk.CTkLabel(
                select_window,
                text="请点击行号选择表头所在行",
                font=("微软雅黑", 12)
            )
            info_label.pack(pady=10)

        except Exception as e:
            messagebox.showerror("错误", f"加载文件失败: {str(e)}")

    def _set_header_row(self, row_num, window):
        self.header_row_var.set(str(row_num))

        try:
            wb = openpyxl.load_workbook(self.baseline_file, data_only=True)
            ws = wb.active

            cols_data = []
            max_col = min(6, ws.max_column)
            for col in range(1, max_col + 1):
                cell_value = ws.cell(row=row_num, column=col).value
                cols_data.append(f"列{col}={str(cell_value) if cell_value else '空'}")

            self.header_preview_label.configure(
                text=f"已选择表头行 {row_num}，内容预览: {', '.join(cols_data)}"
            )
        except Exception as e:
            self.header_preview_label.configure(
                text=f"已选择表头行 {row_num}"
            )

        window.destroy()

    def _select_feature_columns(self):
        if not self.baseline_file:
            messagebox.showerror("错误", "请先选择基准文件")
            return

        try:
            wb = openpyxl.load_workbook(self.baseline_file, data_only=True)
            ws = wb.active

            try:
                header_row = int(self.header_row_var.get())
            except ValueError:
                messagebox.showerror("错误", "表头行号必须是数字")
                return

            max_col = ws.max_column
            header_values = []
            col_name_map = {}
            for col in range(1, max_col + 1):
                cell_value = ws.cell(row=header_row, column=col).value
                col_name = cell_text(cell_value) or "空"
                header_values.append(f"{col}: {col_name}")
                col_name_map[col] = col_name

            select_window = ctk.CTkToplevel(self)
            select_window.title("选择特征列")
            select_window.geometry("400x300")
            select_window.resizable(False, False)

            try:
                icon_path = bundled_path("assets/app.ico")
                if os.path.exists(icon_path):
                    select_window.after(300, lambda: select_window.wm_iconbitmap(icon_path))
            except Exception as e:
                print(f"设置特征列选择窗口图标失败: {e}")

            select_window.transient(self)
            select_window.grab_set()

            listbox = ctk.CTkScrollableFrame(select_window)
            listbox.pack(fill="both", expand=True, padx=10, pady=10)

            checkboxes = []
            for i, header in enumerate(header_values[:20]):
                var = ctk.IntVar()
                checkbox = ctk.CTkCheckBox(listbox, text=header, variable=var)
                checkbox.pack(anchor="w", pady=5)
                checkboxes.append((var, i + 1))

            def on_select():
                selected = [col for var, col in checkboxes if var.get() == 1]
                if len(selected) == 0:
                    messagebox.showerror("错误", "请至少选择1列")
                    return
                if len(selected) > 6:
                    messagebox.showerror("错误", "最多只能选择6列")
                    return

                selected_str = ", ".join(map(str, selected))
                self.feature_cols_var.set(selected_str)

                selected_col_names = [f"{col}({col_name_map[col]})" for col in selected]
                preview_text = f"已选择特征列: {', '.join(selected_col_names)}"

                self.feature_cols_preview_label.configure(
                    text=preview_text
                )

                select_window.destroy()

            select_button = ctk.CTkButton(select_window, text="确定", command=on_select, fg_color="#4CAF50")
            select_button.pack(pady=10)

        except Exception as e:
            messagebox.showerror("错误", f"加载文件失败: {str(e)}")

    def _compare_worker(self):
        try:
            log_queue.put("正在清空results文件夹...")
            for filename in os.listdir(self.results_folder):
                file_path = os.path.join(self.results_folder, filename)
                try:
                    if os.path.isfile(file_path):
                        if os.name == 'nt':
                            subprocess.run(['attrib', '-r', file_path], check=False)
                        else:
                            file_stat = os.stat(file_path)
                            os.chmod(file_path, file_stat.st_mode | stat.S_IWUSR | stat.S_IWGRP | stat.S_IWOTH)
                        os.remove(file_path)
                except Exception as e:
                    log_queue.put(f"删除文件 {filename} 失败: {e}")
            log_queue.put("results文件夹已清空")

            log_queue.put(f"\n已选择基准文件 {self.baseline_file}")
            log_queue.put(f"\n已选择比较文件 {self.compare_file}")

            try:
                header_row = int(self.header_row_var.get())
            except ValueError:
                log_queue.put("\n❌ 错误：表头行号必须是数字")
                return False

            feature_cols_str = self.feature_cols_var.get()
            key_fields = None
            try:
                feature_cols = []
                parts = [p.strip() for p in feature_cols_str.split(",")]
                for part in parts:
                    sub_parts = [sp.strip() for sp in part.split() if sp.strip()]
                    for sub_part in sub_parts:
                        if "-" in sub_part:
                            start, end = map(int, sub_part.split("-"))
                            feature_cols.extend(range(start, end + 1))
                        else:
                            feature_cols.append(int(sub_part))
                feature_cols = sorted(list(set(feature_cols)))
                key_fields = [f"列{col}" for col in feature_cols]
            except ValueError:
                log_queue.put("\n❌ 错误：特征列格式无效")
                return False

            baseline_folder = os.path.basename(os.path.dirname(self.baseline_file))
            compare_folder = os.path.basename(os.path.dirname(self.compare_file))
            baseline_filename = os.path.basename(self.baseline_file).replace('.xlsx', '')
            compare_filename = os.path.basename(self.compare_file).replace('.xlsx', '')
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

            header_preview = ""
            try:
                wb = openpyxl.load_workbook(self.baseline_file, data_only=True)
                ws = wb.active
                if header_row <= ws.max_row:
                    max_col = min(6, ws.max_column)
                    header_cells = []
                    for col in range(1, max_col + 1):
                        cell_value = ws.cell(row=header_row, column=col).value
                        if cell_value:
                            header_cells.append(str(cell_value))
                        else:
                            header_cells.append("空")
                    header_preview = ", ".join(header_cells)
                    if ws.max_column > 6:
                        header_preview += f", ... (共{ws.max_column}列)"
            except Exception as e:
                header_preview = "无法读取表头内容"

            log_queue.put("\n已定义比较配置：")
            log_queue.put(f"\n已选择表头行 {header_row}")
            log_queue.put(f"\n已选择特征列：{feature_cols_str}")

            result_baseline = os.path.join(
                self.results_folder,
                f"{baseline_filename}[基准文件]_比较结果_{timestamp}.xlsx"
            )
            result_compare = os.path.join(
                self.results_folder,
                f"{compare_filename}[比较文件]_比较结果_{timestamp}.xlsx"
            )

            success = compare_excel_files(
                self.baseline_file,
                self.compare_file,
                result_baseline,
                result_compare,
                self.results_folder,
                baseline_filename,
                timestamp,
                header_row,
                key_fields,
                self.stop_event
            )

            if success:
                log_queue.put("\n✅ 任务完成！")
            else:
                log_queue.put("\n❌ 任务失败！")
        except Exception as e:
            log_queue.put(f"\n❌ 任务过程中出错: {str(e)}")
        finally:
            self.running = False
            self.start_button.configure(state="normal")
            self.stop_button.configure(state="disabled")

    def _redirect_stdout(self):
        sys.stdout = StdoutRedirector(self.log_text)

    def _listen_queues(self):
        try:
            while not log_queue.empty():
                message = log_queue.get_nowait()
                if not message.endswith('\n'):
                    message += '\n'

                self.log_text.insert(ctk.END, message)

                line_start = "end-2l"
                line_end = "end-1l"

                if "错误" in message or "Error" in message or "ERROR" in message or "出错" in message:
                    self.log_text.tag_add("error", line_start, line_end)
                    self.log_text.tag_config("error", foreground="#FF5252")
                elif "警告" in message or "Warning" in message or "WARNING" in message:
                    self.log_text.tag_add("warning", line_start, line_end)
                    self.log_text.tag_config("warning", foreground="#FF9800")
                elif "取消" in message:
                    self.log_text.tag_add("cancel", line_start, line_end)
                    self.log_text.tag_config("cancel", foreground="#9E9E9E")
                elif "完成" in message or "成功" in message or "完成!" in message:
                    self.log_text.tag_add("success", line_start, line_end)
                    self.log_text.tag_config("success", foreground="#4CAF50")
                elif "开始" in message or "正在" in message:
                    self.log_text.tag_add("process", line_start, line_end)
                    self.log_text.tag_config("process", foreground="#2196F3")
                elif "已标记" in message or "共发现" in message or "生成" in message:
                    self.log_text.tag_add("result", line_start, line_end)
                    self.log_text.tag_config("result", foreground="#9C27B0")
                else:
                    self.log_text.tag_add("normal", line_start, line_end)
                    self.log_text.tag_config("normal", foreground="#424242")

                self.log_text.see(ctk.END)
        except queue.Empty:
            pass
        finally:
            self.after(100, self._listen_queues)


def main():
    app = ExcelCompareGUI()
    app.mainloop()


if __name__ == "__main__":
    main()
